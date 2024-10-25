# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook, load_workbook
from operator import itemgetter


# # 获取当前工作目录
# current_dir = os.getcwd()
# # 获取当前目录的上层目录
# parent_dir = os.path.dirname(current_dir)
# # 指定你要获取的文件夹名称
# folder_name = "test_data"
# # 构建上层目录中指定文件夹的完整路径
# folder_path = os.path.join(parent_dir, folder_name)
# # 新建文件夹的名称
# new_folder_name = "Post-Processing Data"

# # 构建新文件夹的完整路径
# new_folder_path = os.path.join(folder_path, new_folder_name)

# # 打印新文件夹的路径
# print("Path to 'new' folder in parent directory:", new_folder_path)
# os.makedirs(new_folder_path, exist_ok=True)  # 创建文件夹

file_name = 'summary.xlsx'
new_folder_path = 'F:\doc\\UWB项目文件\测试\BHAST\\0726_bhast'
folder_path = 'F:\doc\\UWB项目文件\测试\BHAST\\0726_bhast'
file_path = os.path.join(new_folder_path, file_name)

wb = Workbook()

wb.save(file_path)




def merge_excel_data(folder_path, output_file):
    wb_summary = Workbook()

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            wb = load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # 如果汇总工作簿中没有当前工作表，则创建一个新的工作表
                if sheet_name not in wb_summary.sheetnames:
                    summary_sheet = wb_summary.create_sheet(title=sheet_name)
                    # 将第一个表格的标题行复制到汇总表格中
                    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                        summary_sheet.append(row)
                else:
                    summary_sheet = wb_summary[sheet_name]

                # 将当前工作表中的数据逐行添加到汇总工作表中
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    summary_sheet.append(row)

    # 删除默认创建的工作表
    del wb_summary["Sheet"]

    # 保存汇总数据到输出文件
    wb_summary.save(output_file)
    print(f"Summary data has been saved to {output_file}")


def sort_sheet_by_id(sheet):
    # 获取工作表中的所有数据行，并转换为列表
    data = list(sheet.iter_rows(values_only=True))
    # 根据第一列 id 进行排序
    sorted_data = sorted(data[1:], key=itemgetter(0))
    # 将排序后的数据写回工作表
    for i, row in enumerate(sorted_data, start=2):
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=value)


def sort_excel_sheets_by_id(file_path):
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sort_sheet_by_id(sheet)
    wb.save(file_path)

merge_excel_data(folder_path, file_path)
sort_excel_sheets_by_id(file_path)
