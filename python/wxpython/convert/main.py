# -*- coding: utf-8 -*-
"""
Created on Sun Apr 21 13:53:56 2019

@author: Administrator
"""

import wx
import MyFrame4
import os
import pandas as pd
import openpyxl as xl
def getERPnumber(path) :
    data = pd.read_excel(path)
    list1 = []
    for i in  range(0,data.shape[0],1):
        tup = tup = (data['规格型号'][i],str(data['存货编码'][i]))
        list1.append(tup)
    dict1 = dict(list1)
    
    return dict1 

def getcalibration(path) :
    data = pd.read_excel(path)
    list1 = []
    for i in  range(0,data.shape[0],1):
        tup = tup = (data['ERROR'][i],str(data['CORRECT'][i]))
        list1.append(tup)
    dict1 = dict(list1)
    
    return dict1

def calibrationwork(row,col,sheet,errorlist) :
    
    for i in range(row,sheet.max_row+1,1):     
    
        if  sheet.cell(row=i,column=col).value in errorlist.keys() :
            sheet.cell(row=i,column=col).value=errorlist[sheet.cell(row=i,column=col).value]
        


def transform(row,col,sheet,erpnumber) :
    for i in range(1,15,1) :
        if sheet.cell(row=row,column=i).value=='物料编码' :
            maxcol=  i
            break
        else:
            maxcol= sheet.max_column+1
    for i in range(row+1,sheet.max_row+1,1):
        try:
            if  sheet.cell(row=i,column=col).value in erpnumber.keys() :
                sheet.cell(row=i,column=maxcol).value=erpnumber[sheet.cell(row=i,column=col).value]
            else:
                sheet.cell(row=i,column=maxcol).value='NA'
        except  Exception as e :
            print(e)
                    
                



def xlstoxlsx (path) :
   
    os.chdir(os.path.dirname(path))
    portion = os.path.splitext(path)
    print(portion[1])
    print(portion[1]+'\n')
    if portion[1]=='.xls' :
        newname = portion[0].split('\\')[-1]+'.xlsx'
        os.rename(path.split('\\')[-1],newname)
        print(newname)
        return newname
    else:
        return path.split('\\')[-1]
def PCBBOM(sheet,erpnumber,errorlist) :
       col = 0
       row = 0
       for i in range(1,10,1) :
           if sheet.cell(row=i,column=1).value=='Designator' :
               row =  i
               break    
       print(row)
       for i in range(1,10,1) :
           if sheet.cell(row=row,column=i).value=='commentname' :
               col =  i
               break    
       print(col)
       calibrationwork(row,col,sheet,errorlist)
       transform(row,col,sheet,erpnumber)

def DeviceandMechanicalBOM(sheet,erpnumber,errorlist) :
       col = 0
       row = 0
       for i in range(1,10,1) :
           if sheet.cell(row=i,column=1).value=='日期' :
               row =  i
               break    
       print(row)
       for i in range(1,10,1) :
           if sheet.cell(row=row+1,column=i).value=='物料型号' :
               col =  i
               break    
       print(col)
       calibrationwork(row,col,sheet,errorlist)
       transform(row+1,col,sheet,erpnumber)
    
                  

def convertwork(path,outpath,erpnumber,errorlist):
   wb = xl.load_workbook(path)
   sheet = wb[wb.sheetnames[0]]
   for i in range(1,10,1) :
           if sheet.cell(row=i,column=1).value=='BOM类型' :
               for t in range(1,5,1) :
                   if sheet.cell(row=i,column=t).value=='产品整机BOM' or sheet.cell(row=i,column=t).value=='产品组装结构BOM' :
                       
                       DeviceandMechanicalBOM(sheet,erpnumber,errorlist)
                       wb.save(outpath+'\\'+path.split('\\')[-1])
                       break
                   if sheet.cell(row=i,column=t).value=='PCBA焊接BOM' :
                       PCBBOM(sheet,erpnumber,errorlist)
                       wb.save(outpath+'\\'+path.split('\\')[-1])
                       break
               break

    
   
   
class MyFrame(MyFrame4.MyFrame2):
	
    def erpget( self,event ):
        dlg = wx.FileDialog(self,'ERP',os.getcwd(),style=wx.FD_MULTIPLE)
        if dlg.ShowModal() == wx.ID_OK :
            self.number1 = getERPnumber(dlg.GetPath())
            print(dlg.GetPaths())
            os.chdir(dlg.GetDirectory())
            self.m_textCtrl5.AppendText('dir is '+os.getcwd())
            self.m_textCtrl5.AppendText(dlg.GetPaths()[0]+'\n')
               
        dlg.Destroy()
    def outpathget( self, event ):
        dlg = wx.DirDialog(self,'output',os.getcwd())
        if dlg.ShowModal() == wx.ID_OK :
               
               self.m_textCtrl5.AppendText(dlg.GetPath())
               self.pathout = dlg.GetPath()
        dlg.Destroy()
    def convert( self, event ):
        self.m_textCtrl5.AppendText(self.pathout)
       
        dlg = wx.FileDialog(self,'ERP',os.getcwd(),style=wx.FD_MULTIPLE)
        if dlg.ShowModal() == wx.ID_OK :
            self.m_textCtrl5.AppendText(self.pathout)
            print(dlg.GetPaths())
           
            for path in  dlg.GetPaths() :
                try :
                    convertwork(xlstoxlsx(path),self.pathout,self.number1,self.errorlist)
                except Exception as  e:
                    print(e)
        dlg.Destroy()
        self.m_textCtrl5.AppendText('done')
        
    def calibrationget( self, event ):
        dlg = wx.FileDialog(self,'ERP',os.getcwd(),style=wx.FD_MULTIPLE)
        if dlg.ShowModal() == wx.ID_OK :
            self.errorlist = getcalibration(dlg.GetPath())
            print(dlg.GetPaths())
            os.chdir(dlg.GetDirectory())
            self.m_textCtrl5.AppendText('dir is '+os.getcwd())
            self.m_textCtrl5.AppendText(dlg.GetPaths()[0]+'\n')
         
  
     
     
if __name__== '__main__' :
    
        app = wx.App(redirect=True)
        
        main_win = MyFrame(None)
        main_win.Show()

        app.MainLoop()